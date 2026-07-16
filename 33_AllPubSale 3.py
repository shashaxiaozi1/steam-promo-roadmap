#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Steam Publisher Sale Builder

功能：
1. 爬取多个 Steam Publisher / Group 的 News
2. 自动判断哪些 News 是促销 Sale
3. 自动整理 Sale Name / Start Date / End Date / Games / URL
4. 只保留 end_date >= 2025-01-01 的促销
5. 支持 manual_overrides.json 人工修正 start_date / end_date
6. 输出：
   - steam_sales_output/promo_data.json
   - steam_sales_output/publisher_sales.xlsx
   - steam_sales_output/latest_update.json
   - steam_sales_output/raw_news/*.json

运行前安装：
pip install requests pandas openpyxl
"""

import json
import re
import time
import html
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple
from urllib.parse import urlencode, unquote
from datetime import datetime, timezone, timedelta


# =========================
# 基本设置
# =========================

COMPANIES = [
    {
        "name": "SEGA",
        "slug": "sega",
        "group_id": 32528477,
        "url": "https://store.steampowered.com/news/group/32528477",
    },
    {
        "name": "Square Enix",
        "slug": "square_enix",
        "group_id": 1012195,
        "url": "https://store.steampowered.com/news/group/1012195",
    },
    {
        "name": "Bandai Namco",
        "slug": "bandai_namco",
        "group_id": 33042543,
        "url": "https://store.steampowered.com/news/group/33042543",
    },
    {
        "name": "Koei Tecmo",
        "slug": "koei_tecmo",
        "group_id": 33016879,
        "url": "https://store.steampowered.com/news/group/33016879",
    },
    {
        "name": "Konami",
        "slug": "konami",
        "group_id": 39026134,
        "url": "https://store.steampowered.com/news/group/39026134",
    },
    {
        "name": "Capcom",
        "slug": "capcom",
        "group_id": 33273264,
        "url": "https://store.steampowered.com/news/group/33273264",
    },
]

API_URL = "https://store.steampowered.com/events/ajaxgetpartnereventspageable/"
LANGUAGE = "english"

# None = 全量爬取。测试时可以改成 200。
MAX_ITEMS_PER_COMPANY: Optional[int] = None

PAGE_COUNT = 100
MAX_PAGES = 1000
SLEEP_SECONDS = 0.5

OUTPUT_ROOT = Path("steam_sales_output")
RAW_NEWS_DIR = OUTPUT_ROOT / "raw_news"
PROMO_JSON = OUTPUT_ROOT / "promo_data.json"
EXCEL_OUTPUT = OUTPUT_ROOT / "publisher_sales.xlsx"
LATEST_UPDATE_JSON = OUTPUT_ROOT / "latest_update.json"

# 只保留 2025-01-01 以后仍在进行/结束的促销
MIN_END_DATE = "2025-01-01"

# 人工修正文件，放在 repo 根目录
MANUAL_OVERRIDES_JSON = Path("manual_overrides.json")


# =========================
# Sale 判断规则
# =========================

STRONG_TITLE_KEYWORDS = [
    "sale",
    "publisher sale",
    "franchise sale",
    "weekend deal",
    "daily deal",
    "midweek deal",
    "special promotion",
    "discount",
    "discounts",
]

BODY_SALE_KEYWORDS = [
    "on sale",
    "now on sale",
    "is on sale",
    "are on sale",
    "save up to",
    "up to",
    "% off",
    "discounted",
    "discounts",
    "limited time",
    "limited-time",
    "offer ends",
    "ends on",
    "until",
    "steam sale",
    "grab",
    "deal",
    "deals",
]

NEGATIVE_KEYWORDS = [
    "free update",
    "patch notes",
    "hotfix",
    "maintenance",
    "trailer",
    "soundtrack release",
    "developer diary",
    "dev diary",
    "behind the scenes",
    "release notes",
]


# =========================
# 通用处理
# =========================

def ensure_dirs() -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    RAW_NEWS_DIR.mkdir(parents=True, exist_ok=True)


def build_detail_url(group_id: int, event_id: str) -> str:
    return f"https://store.steampowered.com/news/group/{group_id}/view/{event_id}?l={LANGUAGE}"


def clean_text(text: str) -> str:
    if not text:
        return ""

    text = html.unescape(str(text))
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("\xa0", " ")

    text = re.sub(r"\[/?list\]", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"\[\*\]", "\n- ", text)

    text = re.sub(r"\[(h1|h2|h3|h4|b|i|u|strike)\]", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\[/(h1|h2|h3|h4|b|i|u|strike)\]", "", text, flags=re.IGNORECASE)

    text = re.sub(
        r"\[url=[^\]]+\](.*?)\[/url\]",
        r"\1",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )

    text = re.sub(
        r"\[url\](.*?)\[/url\]",
        r"\1",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )

    text = re.sub(r"\[img\].*?\[/img\]", "", text, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(
        r"\[previewyoutube=.*?\].*?\[/previewyoutube\]",
        "",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    text = re.sub(r"\[video.*?\].*?\[/video\]", "", text, flags=re.IGNORECASE | re.DOTALL)

    text = re.sub(r"\[store\](.*?)\[/store\]", r"\1", text, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(r"\[app\](.*?)\[/app\]", r"\1", text, flags=re.IGNORECASE | re.DOTALL)

    text = re.sub(r"\[[^\]]+\]", "", text)
    text = re.sub(r"[ \t]+", " ", text)

    lines = []
    for line in text.split("\n"):
        line = line.strip()
        if line:
            lines.append(line)

    return "\n".join(lines).strip()


def parse_steam_ts(value: Any) -> Tuple[str, str]:
    """
    返回：
    - datetime string: YYYY-MM-DD HH:MM:SS
    - date string: YYYY-MM-DD

    使用 UTC，避免本机 / GitHub Actions 时区不同导致日期差一天。
    """
    if value is None or value == "":
        return "", ""

    try:
        ts = int(value)
        dt = datetime.fromtimestamp(ts, tz=timezone.utc)
        return dt.strftime("%Y-%m-%d %H:%M:%S"), dt.strftime("%Y-%m-%d")
    except Exception:
        return str(value), str(value)


def get_now_jst_string() -> str:
    jst = timezone(timedelta(hours=9))
    return datetime.now(tz=jst).strftime("%Y-%m-%d %H:%M:%S JST")


# =========================
# Steam API 爬取
# =========================

def get_events_page(group_id: int, offset: int, count: int) -> Dict[str, Any]:
    import requests

    params = {
        "clan_accountid": group_id,
        "appid": 0,
        "offset": offset,
        "count": count,
        "l": LANGUAGE,
    }

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Referer": f"https://store.steampowered.com/news/group/{group_id}?l={LANGUAGE}",
        "X-Requested-With": "XMLHttpRequest",
    }

    print(f"Request group_id={group_id}, offset={offset}, count={count}")
    print("  " + API_URL + "?" + urlencode(params))

    response = requests.get(API_URL, params=params, headers=headers, timeout=30)
    response.raise_for_status()
    return response.json()


def normalize_event(
    item: Dict[str, Any],
    index: int,
    company: Dict[str, Any],
) -> Dict[str, Any]:
    announcement_body = item.get("announcement_body") or {}

    event_id = str(item.get("gid", "") or item.get("announcement_gid", "") or "")
    raw_body = announcement_body.get("body") or item.get("body") or item.get("summary") or ""

    title = (
        item.get("event_name")
        or announcement_body.get("headline")
        or item.get("headline")
        or ""
    )

    subtitle = (
        item.get("event_notes")
        or item.get("subtitle")
        or announcement_body.get("subtitle")
        or ""
    )

    start_raw = item.get("rtime32_start_time")
    end_raw = item.get("rtime32_end_time")
    start_time, start_date = parse_steam_ts(start_raw)
    end_time, end_date = parse_steam_ts(end_raw)

    posttime_raw = (
        announcement_body.get("posttime")
        or item.get("rtime32_start_time")
        or item.get("rtime32_last_modified")
    )

    updatetime_raw = (
        announcement_body.get("updatetime")
        or item.get("rtime32_last_modified")
    )

    return {
        "index": index,
        "publisher": company["name"],
        "publisher_slug": company["slug"],
        "group_id": company["group_id"],
        "event_id": event_id,
        "app_id": item.get("appid") or announcement_body.get("appid") or "",
        "source_title": clean_text(title),
        "subtitle": clean_text(subtitle),
        "event_type": item.get("event_type"),
        "event_type_name": item.get("event_type_name") or "",
        "start_time_raw": "" if start_raw is None else str(start_raw),
        "end_time_raw": "" if end_raw is None else str(end_raw),
        "start_time": start_time,
        "end_time": end_time,
        "start_date": start_date,
        "end_date": end_date,
        "posttime_raw": "" if posttime_raw is None else str(posttime_raw),
        "updatetime_raw": "" if updatetime_raw is None else str(updatetime_raw),
        "url": build_detail_url(company["group_id"], event_id) if event_id else "",
        "body_text": clean_text(raw_body),
        "body_raw": raw_body,
        "raw": item,
    }


def crawl_company(company: Dict[str, Any]) -> List[Dict[str, Any]]:
    all_items: List[Dict[str, Any]] = []
    seen_ids = set()
    offset = 0

    print("\n" + "=" * 100)
    print(f"Start crawling: {company['name']} / {company['url']}")
    print("=" * 100)

    for page_no in range(1, MAX_PAGES + 1):
        data = get_events_page(company["group_id"], offset, PAGE_COUNT)
        events = data.get("events") or []

        print(f"Page {page_no}: got {len(events)} events")

        if not events:
            break

        new_count = 0

        for item in events:
            event_id = str(item.get("gid", "") or item.get("announcement_gid", "") or "")

            if not event_id or event_id in seen_ids:
                continue

            seen_ids.add(event_id)

            normalized = normalize_event(
                item=item,
                index=len(all_items) + 1,
                company=company,
            )

            all_items.append(normalized)
            new_count += 1

            print(f"[{len(all_items)}] {normalized['source_title']}")

            if MAX_ITEMS_PER_COMPANY is not None and len(all_items) >= MAX_ITEMS_PER_COMPANY:
                save_raw_company_news(company, all_items)
                return all_items[:MAX_ITEMS_PER_COMPANY]

        if new_count == 0:
            print("No new events. Stop.")
            break

        offset += PAGE_COUNT
        time.sleep(SLEEP_SECONDS)

    save_raw_company_news(company, all_items)
    return all_items


def save_raw_company_news(company: Dict[str, Any], items: List[Dict[str, Any]]) -> None:
    path = RAW_NEWS_DIR / f"{company['slug']}_news.json"

    with path.open("w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)

    print(f"Saved raw news: {path}")


# =========================
# Sale 解析
# =========================

def keyword_hits(text: str, keywords: List[str]) -> List[str]:
    t = text.lower()
    return [kw for kw in keywords if kw.lower() in t]


def has_discount_pattern(text: str) -> bool:
    t = text.lower()

    patterns = [
        r"\b\d{1,3}\s*%\s*off\b",
        r"\bsave\s+up\s+to\s+\d{1,3}\s*%",
        r"\bup\s+to\s+\d{1,3}\s*%",
        r"\b\d{1,3}\s*%\s+discount\b",
    ]

    return any(re.search(p, t) for p in patterns)


def classify_sale(event: Dict[str, Any]) -> Dict[str, Any]:
    title = event.get("source_title", "") or ""
    subtitle = event.get("subtitle", "") or ""
    body = event.get("body_text", "") or ""
    event_type_name = event.get("event_type_name", "") or ""

    title_text = f"{title} {subtitle}".strip()
    all_text = f"{title_text}\n{event_type_name}\n{body}".strip()

    reasons = []
    score = 0.0

    strong_title_hits = keyword_hits(title_text, STRONG_TITLE_KEYWORDS)
    if strong_title_hits:
        score += 0.65
        reasons.append("title: " + ", ".join(strong_title_hits[:5]))

    body_hits = keyword_hits(body, BODY_SALE_KEYWORDS)
    if body_hits:
        score += 0.25
        reasons.append("body: " + ", ".join(body_hits[:5]))

    event_type_hits = keyword_hits(event_type_name, ["sale", "discount", "promotion"])
    if event_type_hits:
        score += 0.25
        reasons.append("event_type: " + ", ".join(event_type_hits))

    if has_discount_pattern(all_text):
        score += 0.35
        reasons.append("discount pattern")

    negative_hits = keyword_hits(title_text, NEGATIVE_KEYWORDS)
    if negative_hits:
        score -= 0.35
        reasons.append("negative title: " + ", ".join(negative_hits[:5]))

    if event.get("start_date") and event.get("end_date"):
        score += 0.05
        reasons.append("has date range")

    score = max(0.0, min(1.0, score))
    is_sale = score >= 0.60

    return {
        "is_sale": is_sale,
        "sale_confidence": round(score, 2),
        "sale_reason": " | ".join(reasons) if reasons else "",
    }


def normalize_sale_name(event: Dict[str, Any]) -> str:
    title = event.get("source_title", "").strip()

    title = re.sub(r"\s*[-–—|]\s*Steam.*$", "", title, flags=re.IGNORECASE)
    title = re.sub(r"\s+is\s+live!?$", "", title, flags=re.IGNORECASE)
    title = re.sub(r"\s+starts\s+now!?$", "", title, flags=re.IGNORECASE)
    title = re.sub(r"\s+has\s+started!?$", "", title, flags=re.IGNORECASE)
    title = re.sub(r"\s+now\s+on\s+sale!?$", "", title, flags=re.IGNORECASE)

    if not title:
        title = event.get("subtitle", "").strip()

    return title or "Untitled Sale"


def extract_games_from_body(body_raw: str, body_text: str) -> List[str]:
    games = []
    raw = body_raw or ""

    url_title_matches = re.findall(
        r"\[url=(?:https?:\/\/)?store\.steampowered\.com\/(?:app|sub|bundle)\/[^\]]+\](.*?)\[/url\]",
        raw,
        flags=re.IGNORECASE | re.DOTALL,
    )

    for name in url_title_matches:
        name = clean_text(name).strip()
        if name and len(name) <= 120:
            games.append(name)

    url_slug_matches = re.findall(
        r"store\.steampowered\.com\/(?:app|sub|bundle)\/\d+\/([^\/\]\s?]+)",
        raw,
        flags=re.IGNORECASE,
    )

    for slug in url_slug_matches:
        name = unquote(slug).replace("_", " ").strip()
        if name and len(name) <= 120:
            games.append(name)

    for line in (body_text or "").splitlines():
        line = line.strip()

        if not line.startswith("- "):
            continue

        candidate = line[2:].strip()
        candidate = re.sub(r"\s*[-–—]\s*\d{1,3}%.*$", "", candidate)
        candidate = re.sub(r"\s*\(\s*\d{1,3}%.*?\)\s*$", "", candidate)
        candidate = candidate.strip(" -–—•\t")

        if not candidate:
            continue

        lower = candidate.lower()

        if any(x in lower for x in [
            "wishlist",
            "follow us",
            "join us",
            "sale ends",
            "click here",
            "learn more",
            "check out",
        ]):
            continue

        if 2 <= len(candidate) <= 120:
            games.append(candidate)

    seen = set()
    unique_games = []

    for game in games:
        key = game.lower()

        if key not in seen:
            seen.add(key)
            unique_games.append(game)

    return unique_games


def event_to_promo(event: Dict[str, Any]) -> Dict[str, Any]:
    classification = classify_sale(event)
    games = extract_games_from_body(
        event.get("body_raw", ""),
        event.get("body_text", ""),
    )

    return {
        "publisher": event["publisher"],
        "publisher_slug": event["publisher_slug"],
        "sale_name": normalize_sale_name(event),
        "start_date": event.get("start_date", ""),
        "end_date": event.get("end_date", ""),
        "start_time": event.get("start_time", ""),
        "end_time": event.get("end_time", ""),
        "start_date_raw": event.get("start_time_raw", ""),
        "end_date_raw": event.get("end_time_raw", ""),
        "games": games,
        "games_text": "\n".join(games),
        "source_title": event.get("source_title", ""),
        "source_url": event.get("url", ""),
        "event_id": event.get("event_id", ""),
        "event_type": event.get("event_type", ""),
        "event_type_name": event.get("event_type_name", ""),
        "is_sale": classification["is_sale"],
        "sale_confidence": classification["sale_confidence"],
        "sale_reason": classification["sale_reason"],
        "manual_override_applied": False,
        "body_text": event.get("body_text", ""),
    }


# =========================
# 人工修正 / 更新差分
# =========================

def promo_identity(promo: Dict[str, Any]) -> str:
    event_id = str(promo.get("event_id", "") or "").strip()

    if event_id:
        return f"event_id::{event_id}"

    return "fallback::" + "||".join([
        str(promo.get("publisher", "") or "").strip(),
        str(promo.get("sale_name", "") or "").strip(),
        str(promo.get("start_date", "") or "").strip(),
        str(promo.get("end_date", "") or "").strip(),
    ])


def load_existing_promos() -> List[Dict[str, Any]]:
    if not PROMO_JSON.exists():
        return []

    try:
        with PROMO_JSON.open("r", encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, list):
            return data

        return []
    except Exception as e:
        print(f"Failed to load existing promo_data.json: {repr(e)}")
        return []


def load_manual_overrides() -> Dict[str, Any]:
    if not MANUAL_OVERRIDES_JSON.exists():
        template = {
            "_comment": "Use Steam event_id as key. Only start_date and end_date are supported for now.",
            "_example": {
                "1234567890": {
                    "start_date": "2026-02-13",
                    "end_date": "2026-02-27"
                }
            }
        }

        with MANUAL_OVERRIDES_JSON.open("w", encoding="utf-8") as f:
            json.dump(template, f, ensure_ascii=False, indent=2)

        print(f"Created empty manual overrides file: {MANUAL_OVERRIDES_JSON}")
        return {}

    try:
        with MANUAL_OVERRIDES_JSON.open("r", encoding="utf-8") as f:
            data = json.load(f)

        if not isinstance(data, dict):
            return {}

        return {
            str(k): v
            for k, v in data.items()
            if not str(k).startswith("_") and isinstance(v, dict)
        }

    except Exception as e:
        print(f"Failed to load manual_overrides.json: {repr(e)}")
        return {}


def apply_manual_overrides(promos: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    overrides = load_manual_overrides()

    if not overrides:
        return promos

    fixed_sales_count = 0

    for promo in promos:
        event_id = str(promo.get("event_id", "") or "").strip()

        if not event_id:
            continue

        if event_id not in overrides:
            continue

        rule = overrides[event_id]
        changed = False

        if rule.get("start_date"):
            promo["start_date"] = str(rule["start_date"])
            changed = True

        if rule.get("end_date"):
            promo["end_date"] = str(rule["end_date"])
            changed = True

        if changed:
            promo["manual_override_applied"] = True
            fixed_sales_count += 1

    print(f"Manual override sales applied: {fixed_sales_count}")

    return promos


def filter_promos_by_min_end_date(promos: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    filtered = []

    for promo in promos:
        end_date = str(promo.get("end_date", "") or "").strip()

        if not end_date:
            continue

        if end_date >= MIN_END_DATE:
            filtered.append(promo)

    return filtered


def build_latest_update(
    old_promos: List[Dict[str, Any]],
    new_promos: List[Dict[str, Any]],
) -> Dict[str, Any]:
    old_keys = {promo_identity(p) for p in old_promos}

    new_sales = [
        p for p in new_promos
        if promo_identity(p) not in old_keys
    ]

    return {
        "updated_at": get_now_jst_string(),
        "min_end_date": MIN_END_DATE,
        "total_sales": len(new_promos),
        "new_sales_count": len(new_sales),
        "new_sales": [
            {
                "publisher": p.get("publisher", ""),
                "sale_name": p.get("sale_name", ""),
                "start_date": p.get("start_date", ""),
                "end_date": p.get("end_date", ""),
                "event_id": p.get("event_id", ""),
                "source_url": p.get("source_url", ""),
            }
            for p in new_sales
        ],
    }


# =========================
# 输出
# =========================

def save_promo_json(promos: List[Dict[str, Any]]) -> None:
    with PROMO_JSON.open("w", encoding="utf-8") as f:
        json.dump(promos, f, ensure_ascii=False, indent=2)

    print(f"Saved promo JSON: {PROMO_JSON}")


def save_latest_update(latest_update: Dict[str, Any]) -> None:
    with LATEST_UPDATE_JSON.open("w", encoding="utf-8") as f:
        json.dump(latest_update, f, ensure_ascii=False, indent=2)

    print(f"Saved latest update JSON: {LATEST_UPDATE_JSON}")


def save_excel(promos: List[Dict[str, Any]], all_review_rows: List[Dict[str, Any]]) -> None:
    try:
        import pandas as pd
    except ImportError:
        print("pandas not installed. Skip Excel output.")
        print("Install with: pip install pandas openpyxl")
        return

    promo_rows = []

    for p in promos:
        promo_rows.append({
            "Publisher": p["publisher"],
            "Sale Name": p["sale_name"],
            "Start Date": p["start_date"],
            "End Date": p["end_date"],
            "Games": p["games_text"],
            "Confidence": p["sale_confidence"],
            "Reason": p["sale_reason"],
            "Manual Override Applied": p.get("manual_override_applied", False),
            "Source Title": p["source_title"],
            "URL": p["source_url"],
            "Event Type": p["event_type_name"],
            "Event ID": p["event_id"],
            "Start Raw": p["start_date_raw"],
            "End Raw": p["end_date_raw"],
        })

    review_rows = []

    for p in all_review_rows:
        review_rows.append({
            "Publisher": p["publisher"],
            "Is Sale": p["is_sale"],
            "Confidence": p["sale_confidence"],
            "Reason": p["sale_reason"],
            "Sale Name": p["sale_name"],
            "Start Date": p["start_date"],
            "End Date": p["end_date"],
            "Games": p["games_text"],
            "Source Title": p["source_title"],
            "URL": p["source_url"],
            "Event Type": p["event_type_name"],
            "Event ID": p["event_id"],
        })

    promo_df = pd.DataFrame(promo_rows)
    review_df = pd.DataFrame(review_rows)

    with pd.ExcelWriter(EXCEL_OUTPUT, engine="openpyxl") as writer:
        promo_df.to_excel(writer, sheet_name="Promo List", index=False)
        review_df.to_excel(writer, sheet_name="Review All News", index=False)

        if not promo_df.empty:
            for publisher, sub in promo_df.groupby("Publisher"):
                sheet_name = re.sub(r"[\[\]\:\*\?\/\\]", "_", publisher)[:31]
                sub.to_excel(writer, sheet_name=sheet_name, index=False)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = load_workbook(EXCEL_OUTPUT)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            header_fill = PatternFill("solid", fgColor="D9EAF7")
            header_font = Font(bold=True, color="1F2937")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for col_idx, column_cells in enumerate(ws.columns, start=1):
                max_len = 10
                col_letter = get_column_letter(col_idx)

                for cell in list(column_cells)[:100]:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(value), 60))

                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "\n" in cell.value:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

        wb.save(EXCEL_OUTPUT)

    except Exception as e:
        print(f"Excel styling skipped: {repr(e)}")

    print(f"Saved Excel: {EXCEL_OUTPUT}")


# =========================
# 主程序
# =========================

def main() -> None:
    ensure_dirs()

    old_promos = load_existing_promos()
    all_events: List[Dict[str, Any]] = []

    for company in COMPANIES:
        try:
            items = crawl_company(company)
            all_events.extend(items)
            time.sleep(SLEEP_SECONDS)

        except Exception as e:
            print("\n" + "=" * 100)
            print(f"ERROR while crawling {company['name']}")
            print(f"Error: {repr(e)}")
            print("=" * 100)
            continue

    print("\n" + "=" * 100)
    print(f"Total raw news collected: {len(all_events)}")

    review_rows = [event_to_promo(e) for e in all_events]
    promos = [p for p in review_rows if p["is_sale"]]

    print(f"Detected sale news before filter: {len(promos)}")

    promos = apply_manual_overrides(promos)
    promos = filter_promos_by_min_end_date(promos)

    promos.sort(
        key=lambda x: (
            x.get("publisher", ""),
            x.get("start_date", ""),
            x.get("sale_name", ""),
        )
    )

    review_rows.sort(
        key=lambda x: (
            x.get("publisher", ""),
            -float(x.get("sale_confidence", 0)),
            x.get("source_title", ""),
        )
    )

    latest_update = build_latest_update(old_promos, promos)

    print(f"Detected sale news after date filter: {len(promos)}")
    print(f"New sales this update: {latest_update['new_sales_count']}")

    save_promo_json(promos)
    save_excel(promos, review_rows)
    save_latest_update(latest_update)

    print("\nDone.")
    print(f"Output folder: {OUTPUT_ROOT.resolve()}")
    print("Please check:")
    print(f"  1) {PROMO_JSON}")
    print(f"  2) {EXCEL_OUTPUT}")
    print(f"  3) {LATEST_UPDATE_JSON}")
    print(f"  4) {MANUAL_OVERRIDES_JSON}")


if __name__ == "__main__":
    main()
